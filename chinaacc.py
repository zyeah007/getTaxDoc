#!/usr/bin/env python
# coding=utf-8

'''
从中华会计网校网站的财经法规网页爬取特定名称、地区、年份的法规文档html地址列表，保存至excel文件中。
本代码包含一个类docSpider，和两个函数crawlAll及update。实际使用过程直接调用crawlAll或者update函数。
docSpider代码用于解析网页并保存信息
crawlAll函数会创建一个新的法规列表文件，爬取某类法规项下全部法规列表
update函数更新文件夹目录中已有的某个文件内已爬取的法规列表。
'''
import re
from bs4 import BeautifulSoup as bs
import time
from urllib.parse import urljoin
import requests
from random import randrange
import sys
from pandas import DataFrame
import pandas as pd
import os
import openpyxl


class docSpider(object):
    def __init__(self, tax='', area='', year=''):
        self.start_url = 'http://www.chinaacc.com/fagui/search.shtm'  # 该url地址指向财经法规网页
        self.tax = tax
        self.area = area
        self.year = year
        self.headers = ['标题', '发文单位', '文号', '发文日期', '链接']

    def getToBasePage(self):
        '''
        进入网站的"税收法规"页面
        :return:
        '''
        try:
            r = requests.get(self.start_url)
            r.raise_for_status()
            r.encoding = r.apparent_encoding
            print('主页访问成功')
            return r.text
        except:
            print('访问主页失败')
            return

    def getIDList(self):
        '''
        访问主页，获取税种目录和地区的id编号
        :return: (category_id_dict, area_id_dict)
        '''
        categoryIDs = {'': '', 'name': 'category'}
        areaIDs = {'': '', 'name': 'area'}
        html = self.getToBasePage()
        soup = bs(html, 'html.parser')
        for tag in [categoryIDs, areaIDs]:
            span_list = soup.find_all('div', {'id': tag['name']})[0]('span')
            for c in span_list:
                tag[c.string] = c.attrs['id']
        return categoryIDs, areaIDs

    def genLinkFragment(self, tax, area, year):
        '''
        根据类别、地区和年份的编号组成url地址中的字符串
        :param tax: 要爬取的税种名称
        :return: 目标url中的参数部分
        '''
        categoryIDs, areaIDs = self.getIDList()
        category_id = categoryIDs[tax]
        area_id = areaIDs[area]
        temp = ''
        if category_id:
            temp += '&category=%s' % category_id
        if area_id:
            temp += '&area=%s' % area_id
        if year:
            temp += '&time=%s' % year
        return temp

    def parseHTML(self, html):
        '''

        :param html:根据税种、地区、年份确定的URL链接的首页，由该页面生成html文件
        :return:(df, nextPage)，由该页面要爬取的信息生成的DataFrame，以及下一页的链接
        '''
        soup = bs(html, 'html.parser')
        result = soup.find(class_='lhnr clearfix')
        docs = result.find_all(class_='nr clearfix')
        pattern = r'(发文单位.+?)?(文号.+?)?(发文日期.+)'
        out_list = []
        nextPage = ''
        for t in docs:
            info = {}
            link = urljoin(self.start_url, t.a['href'])
            title = t.a['title'].strip()
            infos = t.find(class_='c').stripped_strings
            try:
                info_text = ''
                for s in infos:
                    info_text += s
                match = re.search(pattern, info_text)
                pub_org = match.group(1).strip().split('：')[1].strip() if isinstance(match.group(1), str) else ''
                pub_num = match.group(2).strip().split('：')[1].strip() if isinstance(match.group(2), str) else ''
                pub_date = match.group(3).strip().split('：')[1].strip() if isinstance(match.group(3), str) else ''
            except:
                pub_org = ''
                pub_num = ''
                pub_date = ''
            info['标题'] = title
            info['发文单位'] = pub_org
            info['文号'] = pub_num
            info['发文日期'] = pub_date
            info['链接'] = link
            out_list.append(info)
        df = DataFrame(data=out_list, columns=self.headers)
        fy = soup.find(class_='fy clearfix msf')
        for a in fy('a'):
            if a.text == '下一页':  # 只有最后一页没有"下一页"标签，故可由此判断是否存在下一页。
                nextPage = a['href']
                break
            else:
                nextPage = ''
        return df, nextPage

    def loop(self, target_url):
        '''
        解析单一页面信息。本函数与parseHTML函数返回数据相同，只是担心偶尔访问目标url失败，在函数内设置递归
        :param target_url:
        :return:
        '''
        try:
            r = requests.get(target_url)
            r.raise_for_status()
            r.encoding = r.apparent_encoding
            print('成功访问网页:%s' % target_url)
            df, nextPage = self.parseHTML(r.text)
            return df, nextPage
        except:
            print('访问指定税收法规网页失败:%s' % target_url)
            time.sleep(randrange(6, 10))
            self.loop(target_url)

    def crawl(self, tax='', area='', year=''):
        '''

        :param tax: 要爬取对法规的名称
        :param area: 要爬取的地区
        :param year: 要爬取的年份
        :return: 一个数据列表。列表中的单个元素为一个字典，记录法规条文的标题、发布单位、日期、链接等信息
        '''
        link_fragment = self.genLinkFragment(tax, area, year)
        target_url = self.start_url + '?' + 'page=1' + link_fragment # 生成目标url地址
        outList = []
        pageNum = 1
        infoCount = 0 # 统计爬取信息数量
        while target_url:  # 每循环一次，爬取一页的数据
            df, nextPage = self.loop(target_url)
            outList.append(df)
            infoCount += len(df)
            print('已爬取%d条数据' % infoCount)
            if nextPage:
                target_url = urljoin(self.start_url, nextPage, allow_fragments=False)
                pageNum += 1
                print('继续链接第%d页' % pageNum)
                time.sleep(randrange(3, 6))
            else:
                break
        outData = pd.concat(outList, ignore_index=True)
        outData.drop_duplicates()
        return outData

    def updateData(self, tax, fPath, savePath=''):
        '''
        :param tax:法规名称
        :param fPath: 原数据文件。需要是可以读入DataFrame中的数据格式
        :param savePath: 更新后的数据文件。
        :return: 更新后的数据data，格式为列表，列表中每个元素为字典
        '''
        # 首先检查法规与数据文件是否匹配
        if tax not in fPath:
            print('法规名称与数据文件不匹配！请重新确认输入！')
            time.sleep(5)
            sys.exit(0)
        if savePath == '':
            savePath = fPath
        # 读入数据到DataFrame中
        try:
            df = pd.read_excel(fPath, sheet_name='法规列表', header=0)
        except:
            df = pd.read_excel(fPath,header=0)
        header=df.columns
        # 取得最近一个发布日的时间戳
        df2 = df.dropna(subset=['发文日期'])
        timestamp = df2['发文日期'].max()
        temp_df = df2[df2['发文日期'] == timestamp]  # 取一个临时的数据集，用来跟新爬取的数据记录比较，以删除重复记录
        print('当前文件记录的时间戳为：%s' % timestamp)
        # 开始更新数据
        link_fragment = self.genLinkFragment(tax, '', '')
        target_url = self.start_url + '?' + 'page=1' + link_fragment
        outList = []
        while target_url:
            update_df, nextPage = self.loop(target_url)
            outList.append(update_df)
            # 分析新爬取页面的最早时间戳
            ts_early = update_df['发文日期'].min()
            if ts_early < timestamp:  # 如果当前网页最早的发布日期已经早已数据文件的时间戳，则结束爬取
                break
            else:  # 否则，继续翻页爬取下一页的数据
                if nextPage:
                    target_url = urljoin(self.start_url, nextPage, allow_fragments=False)
                    print('继续访问下一页...')
                    time.sleep(randrange(3, 6))
                else:
                    break
        new_df = pd.concat(outList, ignore_index=True)
        # 下面对新爬取的数据集进行处理：
        # 首先，删除发布日期早于timestamp的记录
        new_df = new_df[new_df['发文日期'] >= timestamp]
        # 然后，将原数据中等于时间戳的记录与new_df合并
        temp_df = pd.concat([new_df, temp_df], ignore_index=True)
        # 最后，删除重复值
        temp_df.drop_duplicates(['标题'], keep='last', inplace=True)
        result = pd.concat([temp_df, df], ignore_index=True)
        result.drop_duplicates(keep='last',inplace=True)
        result=result.loc[:,header]
        print('更新了%d条数据!' % (result.index.size-df.index.size))
        self.writeData(result, savePath)

    def writeData(self, df, out_path):
        '''

        :param data:爬取到的全部数据，为DataFrame格式，包含标题
        :param out_path: 存储文件到路径
        :return: None
        '''
        # 将爬取的数据写入xlsx文件
        print('正在向xlsx文件中写入数据...')
        df.to_excel(out_path, sheet_name='法规列表', index=False)
        print('数据写入完成！')
        # 为每个标题添加超链接
        wb=openpyxl.load_workbook(out_path)
        sheet=wb['法规列表']
        print('添加超链接...')
        for i in range(2,sheet.max_row+1):
            sheet.cell(row=i,column=1).hyperlink=sheet.cell(row=i,column=5).value
        wb.save(out_path)
        wb.close()
        print('数据文件保存路径为：%s' % out_path)


def crawlAll():
    '''
    创建一个新的法规列表文件，爬取某类法规项下全部法规列表。

    :return:
    '''
    taxName = input('输入要爬取的法规名称：\n')
    area = input('请输入地区：\n')
    year = input('请输入年份:\n')
    start_time = time.time()
    spider = docSpider(tax=taxName, area=area, year=year)
    data = spider.crawl(tax=spider.tax, area=spider.area, year=spider.year)
    if spider.year:
        timeRange = spider.year
    else:
        timeRange = '全部'
    if area:
        areaTag = area
    else:
        areaTag = ''
    out_path = r'./%s%s-%s' % (taxName, areaTag, timeRange) + '.xlsx'
    spider.writeData(data, out_path)
    print('共记录%d条数据' % len(data))
    end_time = time.time()
    print('爬取共用时:%.1f分钟.' % ((end_time - start_time) / 60))
    print('done')


def update():
    '''
    由于当前技术限制，不能在程序执行中自定义待更新待法规文件名称，故暂时只能更新文件名为"*-全部.xlsx"的文件
    :return:
    '''
    taxName = input('请输入要更新的法规名称：\n')
    fPath = os.path.join(os.getcwd(), '%s-全部.xlsx' % taxName)
    newFileName = input('请输入新数据保存的文件名（若输入为空，则默认覆盖原数据文件）：\n')
    if newFileName:
        savePath = os.path.join(os.getcwd(), newFileName) + '.xlsx'
    else:
        savePath = fPath
    spider = docSpider(tax=taxName)
    spider.updateData(taxName, fPath, savePath)


if __name__ == '__main__':
    crawlAll()
